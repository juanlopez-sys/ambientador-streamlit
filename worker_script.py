import os
import time
import requests
import unicodedata
import base64
from io import BytesIO
from PIL import Image, ImageOps
from openpyxl import load_workbook

# ==============================
# CONFIGURACIÓN GENERAL
# ==============================
base_URL_para_imagenes = "https://media.falabella.com/falabellaCL/{sku}/w=1500,h=1500,fit=pad.jpg"
carpeta_temporal_imagenes_github = os.path.join(os.environ.get("RUNNER_TEMP", "."), "imagenes_falabella")
Carpeta_reconstruye_desdde_base64 = os.path.join(carpeta_temporal_imagenes_github, "imagen JPG")
carpeta_salida_IA_JPG = os.path.join(carpeta_temporal_imagenes_github, "ambientada")

Máximo_tiempo_de_espera = 15
Maxima_cantidad_de_intentos = 3
Segundos_espera_reintento = 1.5
HTTP_DEFAULT_HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) PythonRequests/2.x"}

# === OpenAI / gpt-image - 1 ===
OpenAI_API_para_editar = "https://api.openai.com/v1/images/edits"
OPENAI_API_KEY_VALUE = os.environ.get("OPENAI_API_KEY", "")
OPENAI_IMAGE_MODEL_NAME = "gpt-image-1"
OPENAI_IMAGE_OUTPUT_SIZE = "1024x1024"

PROMPTMAP_BY_LINEA = {
    "J11": """
First identify what this product is. Then place this product in a realistic, clean, and modern environment,
preserving its original shape, color, and proportions. It must appear in a context that is coherent with
its function. For example, if its use requires interaction with another object (such as a keyboard next to
a computer, a stand next to a TV, or a person only when the product is a vacuum cleaner), apply a 'less is more'
approach. You may include elements that visually reinforce the product's natural use and environment, as long as
the composition remains balanced and professional.

The product must retain its original appearance without alterations or added parts, and if it has a cable, it
should appear naturally connected or in use. The background should represent a modern, organized indoor space such
as an office, bedroom, living room, kitchen, or studio, with warm or natural lighting and soft shadows. The product
must be fully focused and sharp.

Use realistic surfaces such as light wood, marble, ceramic, or smooth painted walls, while keeping the scene
balanced and professional. Avoid excessive saturation or digital reflections. The result should resemble a
professional lifestyle or tech catalog photograph, with coherent lighting, natural shadows, and the product
integrated in a functional and believable way. The goal is to allow a customer who does not know the product
to understand how it would look in a real environment.
""",
    "J13": """
Place this furniture item in a realistic, clean, and modern environment, preserving its original shape,
color, and proportions. The background should adapt to the type of furniture and may represent interiors such
as a living room, dining room, bedroom, or home office, making it evident that the furniture is placed inside the house.
The furniture must not have any objects placed on it.
The environment must be fully focused and sharp, with natural or warm lighting depending on the context.

Walls may vary between beige, gray, soft green, blue, or terracotta tones, always in harmony with the furniture,
and surfaces should be neutral or natural such as light wood, stone, or microcement. Include contextual elements
coherent with the scene, such as plants, lamps, frames, curtains, pots, pergolas, towels, or bathroom accessories,
but following a 'less is more' approach.

The scene should resemble a professional interior design, with real perspective,
coherent shadows, and visual balance between the furniture and its environment, avoiding blurry backgrounds,
artificial reflections, or excessive saturation.
""",
    "J15": """
Place this product in a realistic, clean, and modern environment, preserving its shape, color, and proportions.
Follow a 'less is more' approach, keeping the scene simple, balanced, and visually coherent.
"""
}

PROMPTMAP_BY_MARCA = {
    "MICA": """
Place this furniture item in a realistic, clean, and modern environment, preserving its original shape, color,
and proportions. The background should adapt to the type of furniture and may represent an interior such as a
living room, dining room, bedroom, or home office.
The furniture must not have any objects placed on it.
The environment must be fully focused and sharp, with natural or warm lighting depending on the context, following
a 'less is more' approach.

Walls may vary between beige, gray, soft green, blue, or terracotta tones, in harmony with the furniture, and
surfaces should be neutral or natural such as light wood, stone, or microcement. Include contextual elements such
as plants, lamps, frames, curtains, pots, pergolas, towels, but avoid clutter or distracting
details.

The scene should resemble a professional interior design, with real perspective,
coherent shadows, and color balance. Avoid artificial reflections, blurry backgrounds, or excessive saturation.
Brand guidelines: youthful, practical, minimalistic, contemporary; simple neutral colors; white plain backgrounds;
neutral-tone rugs.
""",
    "BASEMENT HOME": """
Place this furniture item in a realistic, clean, and modern environment, preserving its original shape and proportions.
The background must adapt to the furniture type, representing interiors such as living room, dining room, bedroom, or
home office.
The furniture must not have any objects placed on it.
The scene must be sharp and fully focused, with natural or warm lighting as appropriate.

Walls should use tones like beige, gray, soft green, blue, or terracotta—always in harmony with the furniture.
Surfaces should be neutral or natural such as light wood, stone, or microcement. Include contextual elements
such as plants, lamps, frames, curtains, pots, pergolas, towels, or discreet accessories while following a
'less is more' approach.

The result should resemble a professional interior design, with real
perspective, coherent shadows, and visual balance. Avoid blurry backgrounds, artificial reflections, or
excessive saturation.
Brand guidelines: contemporary, sober, elegant, everyday style; target audience +35; cement or neutral walls;
neutral-tone rugs; minimalist neutral-tone wall art; references: Ferm Living and West Elm.
""",
    "BASEMENT H": """
Place this furniture item in a realistic, clean, and modern setting, preserving its original shape, color, and
proportions. Adapt the background according to the furniture type: interior (living room, dining room, bedroom,
home office).
The furniture must not have any objects placed on it.
The scene must be sharp and fully focused, with natural or warm lighting.

Walls may use beige, gray, soft green, blue, or terracotta tones. Surfaces must be neutral or natural. Add contextual
elements such as plants, lamps, frames, curtains, pots, pergolas, towels, or small accessories—always following a
'less is more' approach.

The final image should look like a professional interior design: real perspective,
consistent shadows, balanced composition, and no blurry backgrounds, artificial reflections, or over-saturation.
Brand guidelines: contemporary, sober, elegant, everyday style; target audience +35; cement or neutral walls; neutral rugs;
minimalist art; references: Ferm Living and West Elm.
""",
    "ROBERTA ALLEN": """
Place this furniture item in a realistic, clean, modern environment, preserving its original shape and proportions.
The background may represent interiors such as living room, dining room, bedroom, or home office.
The furniture must not have any objects placed on it.
The environment must be sharp and fully focused, with natural or warm lighting.

Walls may vary between beige, gray, soft green, blue, or terracotta tones, with neutral or natural surfaces such as
light wood, stone, or microcement. Include contextual elements such as plants, lamps, frames, curtains, pots, pergolas,
towels, or bathroom accessories, while keeping a 'less is more' approach.

The result should resemble a professional lifestyle, interior design session with correct
perspective, coherent shadows, and visual harmony. Avoid blurry backgrounds, artificial reflections, or high saturation.
Brand guidelines: feminine and romantic, classic and elegant; simple neutral colors; textured or molded walls; neutral-tone
rugs; floral accents; large windows; white linen curtains.
"""
}

def upscale_to_square_jpeg(img_bytes: bytes, target: int = 1500, bg=(255, 255, 255)) -> bytes:
    with Image.open(BytesIO(img_bytes)) as im:
        if im.mode in ("RGBA", "P"):
            im = im.convert("RGB")
        fitted = ImageOps.contain(im, (target, target), Image.Resampling.LANCZOS)
        canvas = Image.new("RGB", (target, target), bg)
        x = (target - fitted.width) // 2
        y = (target - fitted.height) // 2
        canvas.paste(fitted, (x, y))
        out = BytesIO()
        canvas.save(out, format="JPEG", quality=95, optimize=True)
        return out.getvalue()

def _norm(input_text: str) -> str:
    if input_text is None:
        return ""
    local_text = str(input_text).strip()
    if not local_text:
        return ""
    local_text = "".join(ch for ch in unicodedata.normalize("NFD", local_text) if unicodedata.category(ch) != "Mn")
    return local_text.upper()

def clasificar_mueble_desde_imagen(img_bytes: bytes, key_brand, key_line, prompt_actual):
    nuevo_prompt = prompt_actual

    if not OPENAI_API_KEY_VALUE:
        print("[CLASIFICACIÓN] Falta OPENAI_API_KEY, se asume 'interior'")
        return nuevo_prompt

    try:
        img_b64 = base64.b64encode(img_bytes).decode("ascii")
        img_data_url = f"data:image/jpeg;base64,{img_b64}"

        url = "https://api.openai.com/v1/chat/completions"
        headers = {
            "Authorization": f"Bearer {OPENAI_API_KEY_VALUE}",
            "Content-Type": "application/json",
        }

        prompt_text = (
            "Analiza la imagen del mueble y responde SOLO con una de estas palabras: "
            "'terraza' si el mueble está pensado principalmente para exterior/terraza/jardín, "
            "o 'interior' si el mueble es para uso dentro de la casa."
        )

        payload = {
            "model": "gpt-4o-mini",
            "messages": [
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt_text},
                        {"type": "image_url", "image_url": {"url": img_data_url}},
                    ],
                }
            ],
            "max_tokens": 5,
        }

        resp = requests.post(url, headers=headers, json=payload, timeout=60)
        resp.raise_for_status()
        data = resp.json()
        raw = data["choices"][0]["message"]["content"].lower()
        print(raw)

        if "terraza" in raw:
            PROMPTMAP_BY_LINEA_TERRAZA = {
                "J13": """
Place this furniture item in a realistic, clean, and modern outdoor terrace located in a garden.
Always preserve its original shape, color, and proportions. The scene must be set exclusively in
an exterior environment: a terrace surrounded by natural vegetation, grass, trees, bushes, or
landscaped garden elements.

The environment must be fully focused and sharp, with natural outdoor lighting resembling a sunny
or slightly cloudy day. Shadows should be soft and consistent.

Vegetation should appear realistic. Decorative elements must follow a “less is more” approach.
The furniture must not have any objects placed on it.
"""
            }

            PROMPTMAP_BY_MARCA_TERRAZA = {
                "MICA": """
Place this furniture item in a realistic, clean, and modern outdoor environment, specifically on a terrace within a garden.
Preserve its original shape, color, and proportions. Use natural outdoor lighting with soft shadows.
The furniture must not have any objects placed on it.
""",
                "BASEMENT HOME": """
Place this furniture item in a realistic, clean, and modern outdoor environment, specifically on a terrace within a garden.
Preserve its original shape and proportions. The furniture must not have any objects placed on it.
""",
                "BASEMENT H": """
Place this furniture item in a realistic, clean, and modern outdoor setting, specifically on a terrace within a garden.
Preserve its original shape, color, and proportions. The furniture must not have any objects placed on it.
""",
                "ROBERTA ALLEN": """
Place this furniture item in a realistic, clean, and modern outdoor environment, specifically on a terrace within a garden.
Preserve its original shape and proportions. The furniture must not have any objects placed on it.
"""
            }

            if key_line in PROMPTMAP_BY_LINEA_TERRAZA:
                nuevo_prompt = PROMPTMAP_BY_LINEA_TERRAZA[key_line]
            if key_brand in PROMPTMAP_BY_MARCA_TERRAZA:
                nuevo_prompt = PROMPTMAP_BY_MARCA_TERRAZA[key_brand]

    except Exception as e:
        print(f"[CLASIFICACIÓN ERROR] {e} -> se asume 'interior'")

    return nuevo_prompt

def leer_filas_desde_excel(ruta_excel, col_sku="SKU", col_marca="marca", col_linea="linea"):
    wb = load_workbook(ruta_excel, read_only=True, data_only=True)
    hoja = wb.active
    encabezados = [str(c.value).strip() if c.value else "" for c in next(hoja.iter_rows(min_row=1, max_row=1))]
    encabezados_norm = [_norm(h) for h in encabezados]
    requeridos = {_norm(col_sku): "sku", _norm(col_marca): "marca", _norm(col_linea): "linea"}
    idx_map = {}
    for i, h in enumerate(encabezados_norm):
        if h in requeridos and requeridos[h] not in idx_map:
            idx_map[requeridos[h]] = i
    faltantes = [k for k in ["sku", "marca", "linea"] if k not in idx_map]
    if faltantes:
        raise ValueError(f"No se encontraron columnas requeridas en el Excel: {faltantes}. Encabezados detectados: {encabezados}")
    filas = []
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        sku_val = fila[idx_map["sku"]] if idx_map["sku"] < len(fila) else None
        marca_val = fila[idx_map["marca"]] if idx_map["marca"] < len(fila) else None
        linea_val = fila[idx_map["linea"]] if idx_map["linea"] < len(fila) else None
        sku_txt = str(sku_val).strip() if sku_val is not None else ""
        marca_txt = str(marca_val).strip() if marca_val is not None else ""
        linea_txt = str(linea_val).strip() if linea_val is not None else ""
        if sku_txt:
            filas.append({"sku": sku_txt, "marca": marca_txt, "linea": linea_txt})
    wb.close()
    return filas

def enviar_a_gpt_image(source_bytes: bytes, prompt_text: str, timeout_sec: int = 300) -> bytes:
    if not OPENAI_API_KEY_VALUE:
        raise RuntimeError("Falta OPENAI_API_KEY en variables de entorno.")
    files = {"image": ("input.jpg", source_bytes, "image/jpeg")}
    data = {"model": OPENAI_IMAGE_MODEL_NAME, "prompt": prompt_text if prompt_text else "", "size": OPENAI_IMAGE_OUTPUT_SIZE}
    headers = {"Authorization": f"Bearer {OPENAI_API_KEY_VALUE}"}
    resp = requests.post(OpenAI_API_para_editar, headers=headers, files=files, data=data, timeout=timeout_sec)
    if resp.status_code != 200:
        raise RuntimeError(f"gpt-image HTTP {resp.status_code}: {resp.text}")
    js = resp.json()
    try:
        item = js["data"][0]
    except Exception as err:
        raise RuntimeError(f"Respuesta inesperada de gpt-image: {js}") from err
    if "url" in item and item["url"]:
        r2 = requests.get(item["url"], timeout=timeout_sec)
        r2.raise_for_status()
        return r2.content
    if "b64_json" in item and item["b64_json"]:
        return base64.b64decode(item["b64_json"])
    raise RuntimeError(f"Formato de respuesta no soportado: {js}")

if __name__ == "__main__":
    ruta_excel_archivo = os.environ.get("RUTA_EXCEL_ARCHIVO")
    if not ruta_excel_archivo or not os.path.exists(ruta_excel_archivo):
        raise RuntimeError(f"No se recibió RUTA_EXCEL_ARCHIVO válida. Valor: {ruta_excel_archivo}")
    print(f"[WORKER] Usando Excel: {ruta_excel_archivo}")

    input_rows = leer_filas_desde_excel(ruta_excel_archivo, col_sku="SKU", col_marca="marca", col_linea="linea")

    for ensure_dir in (carpeta_temporal_imagenes_github, Carpeta_reconstruye_desdde_base64, carpeta_salida_IA_JPG):
        os.makedirs(ensure_dir, exist_ok=True)

    path_metadata_csv = os.path.join(carpeta_temporal_imagenes_github, "metadata.csv")
    if not os.path.exists(path_metadata_csv):
        with open(path_metadata_csv, "w", encoding="utf-8", newline="") as meta_fp:
            meta_fp.write("sku,image_path,marca,linea,prompt,base64_len,rebuilt_path,ambientada_path\n")

    count_total = 0
    count_ok_download = 0

    for row_item in input_rows:
        count_total += 1
        sku_id = str(row_item.get("sku", "")).strip()
        brand_raw = row_item.get("marca", "")
        line_raw = row_item.get("linea", "")

        key_line = _norm(line_raw)
        key_brand = _norm(brand_raw)
        selected_prompt = "Ambienta esta imágen"
        if key_line in PROMPTMAP_BY_LINEA:
            selected_prompt = PROMPTMAP_BY_LINEA[key_line]
        if key_brand in PROMPTMAP_BY_MARCA:
            selected_prompt = PROMPTMAP_BY_MARCA[key_brand]

        image_url = base_URL_para_imagenes.format(sku=sku_id)
        path_download_jpg = os.path.join(carpeta_temporal_imagenes_github, f"{sku_id}.jpg")

        if os.path.exists(path_download_jpg) and os.path.getsize(path_download_jpg) > 0:
            print(f"[OK] Ya existe: {path_download_jpg}")
            count_ok_download += 1
        else:
            flag_success = False
            for attempt_idx in range(1, Maxima_cantidad_de_intentos + 1):
                try:
                    resp = requests.get(image_url, headers=HTTP_DEFAULT_HEADERS, timeout=Máximo_tiempo_de_espera, stream=True)
                    if resp.status_code == 200 and "image" in resp.headers.get("Content-Type", "").lower():
                        with open(path_download_jpg, "wb") as out_fp:
                            for chunk in resp.iter_content(chunk_size=8192):
                                if chunk:
                                    out_fp.write(chunk)
                        print(f"[OK] {sku_id} -> {path_download_jpg}")
                        count_ok_download += 1
                        flag_success = True
                        break
                    elif resp.status_code == 404:
                        print(f"[NO ENCONTRADA] {sku_id} (404) URL: {image_url}")
                        break
                    else:
                        print(f"[ERROR] {sku_id} HTTP {resp.status_code} intento {attempt_idx}/{Maxima_cantidad_de_intentos}")
                except requests.Timeout:
                    print(f"[TIMEOUT] {sku_id} intento {attempt_idx}/{Maxima_cantidad_de_intentos}")
                except requests.RequestException as e:
                    print(f"[EXCEPCIÓN] {sku_id} intento {attempt_idx}/{Maxima_cantidad_de_intentos}: {e}")
                if attempt_idx < Maxima_cantidad_de_intentos:
                    time.sleep(Segundos_espera_reintento)
            if not flag_success:
                print(f"[FALLO] {sku_id} no se pudo descargar")

        out_base64_len = 0
        path_rebuilt_jpg = os.path.join(Carpeta_reconstruye_desdde_base64, f"{sku_id}.jpg")
        source_img_bytes = None
        try:
            if os.path.exists(path_download_jpg) and os.path.getsize(path_download_jpg) > 0:
                with open(path_download_jpg, "rb") as src_fp:
                    source_img_bytes = src_fp.read()
                    b64_text = base64.b64encode(source_img_bytes).decode("ascii")
                out_base64_len = len(b64_text)
                with open(path_rebuilt_jpg, "wb") as rb_fp:
                    rb_fp.write(base64.b64decode(b64_text))
                print(f"[REBUILD] {sku_id} -> {path_rebuilt_jpg}")
            else:
                print(f"[SKIP REBUILD] {sku_id} no tiene archivo fuente para convertir.")
        except Exception as e:
            print(f"[ERROR BASE64/JPG] {sku_id}: {e}")
            source_img_bytes = None

        if key_line == "J13" and source_img_bytes:
            selected_prompt = clasificar_mueble_desde_imagen(source_img_bytes, key_brand, key_line, selected_prompt)
            print(f"[CLASIFICACIÓN SKU {sku_id}] LINEA J13")

        path_ai_ambient_jpg = os.path.join(carpeta_salida_IA_JPG, f"{sku_id}_005.jpg")
        if source_img_bytes:
            try:
                result_ai_bytes = enviar_a_gpt_image(source_img_bytes, selected_prompt)
                final_1500 = upscale_to_square_jpeg(result_ai_bytes, target=1500, bg=(255, 255, 255))
                with Image.open(BytesIO(final_1500)) as im:
                    if im.mode in ("RGBA", "P"):
                        im = im.convert("RGB")
                    im.save(path_ai_ambient_jpg, format="JPEG", quality=95)
                print(f"[IA OK] {sku_id} -> {path_ai_ambient_jpg}")
            except Exception as e:
                print(f"[IA ERROR] {sku_id}: {e}")
        else:
            print(f"[IA SKIP] {sku_id}: sin bytes de imagen para enviar a la IA.")

        with open(path_metadata_csv, "a", encoding="utf-8", newline="") as meta_fp:
            brand_csv = str(brand_raw).replace('"', '""')
            line_csv = str(line_raw).replace('"', '""')
            prompt_csv = str(selected_prompt).replace('"', '""')
            meta_fp.write(f'{sku_id},"{path_download_jpg}","{brand_csv}","{line_csv}","{prompt_csv}",{out_base64_len},"{path_rebuilt_jpg}","{path_ai_ambient_jpg}"\n')

    print(f"\nResumen: {count_ok_download}/{count_total} imagen(es) procesadas. Carpeta: {carpeta_temporal_imagenes_github}")
    status_path = os.path.join(carpeta_temporal_imagenes_github, "status.done")
    try:
        with open(status_path, "w", encoding="utf-8") as f:
            f.write(f"ok_download={count_ok_download}, total={count_total}\n")
    except Exception:
        pass

    import shutil as _sh
    from datetime import datetime as _dt
    try:
        timestamp = _dt.now().strftime("%Y%m%d_%H%M%S")
        zip_base = os.path.join(carpeta_temporal_imagenes_github, f"ambientadas_{timestamp}")
        zip_path = _sh.make_archive(base_name=zip_base, format="zip", root_dir=carpeta_salida_IA_JPG)
        print(f"\nZIP creado con las imágenes ambientadas: {zip_path}")
    except Exception as e:
        print(f"No se pudo crear el ZIP de imágenes ambientadas: {e}")
